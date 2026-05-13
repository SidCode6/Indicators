"""Download and parse the daily VanEck NODE ETF holdings XLSX.

VanEck publishes the prior trading day's holdings around 10:25 AM ET as an
Excel file at a stable URL on vaneck.com. The site gates content behind a
disclaimer cookie; we pre-set it to avoid a redirect loop.

The XLSX filename contains the as-of date: ``NODE_asof_YYYYMMDD.xlsx``.
We use that as the canonical date for the snapshot.
"""

from __future__ import annotations

import io
import re
from dataclasses import dataclass
from datetime import datetime, timezone
from typing import Optional

import openpyxl
import requests


HOLDINGS_PAGE_URL = (
    "https://www.vaneck.com/us/en/investments/onchain-economy-etf-node/holdings/"
)
HOLDINGS_DOWNLOAD_URL = (
    "https://www.vaneck.com/us/en/investments/onchain-economy-etf-node/downloads/holdings/"
)
FUND_DATASET_URL = (
    "https://www.vaneck.com/Main/FundDatasetBlock/Get/"
    "?blockId=351303&pageId=310955&ticker=NODE"
)

USER_AGENT = (
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0 Safari/537.36"
)

# Pre-set disclaimer-accept cookies so the site doesn't bounce us to /disabled-cookies
DISCLAIMER_COOKIES = {
    "ve-country-us": (
        "iso%3Dus%26investortype%3Dretail%26language%3Den"
        "%26disclaimer%3Dtrue%26foreigntax%3Dfalse%26foreigntaxdisclaimer%3Dfalse"
    ),
    "ve-country": "current%3Dus%26previous%3D",
    "sitelanguage": "en",
}

# VanEck's own funds held inside NODE — these are structural exposure decisions
# (BTC/ETH/SOL/quantum sleeves) rather than active stock picks.
VANECK_FUND_TICKERS = {"HODL", "ETHV", "VSOL", "QNTM LN"}


@dataclass
class FetchResult:
    as_of: str                  # YYYY-MM-DD, from the XLSX filename
    source_filename: str        # e.g. NODE_asof_20260512.xlsx
    fetched_at: str             # ISO 8601 UTC
    fund: dict                  # NAV, total_net_assets_usd, etc.
    holdings: list[dict]
    raw_xlsx_bytes: bytes


def _session() -> requests.Session:
    """Build a session with disclaimer cookies and a browser UA."""
    s = requests.Session()
    s.headers.update({
        "User-Agent": USER_AGENT,
        "Accept": (
            "text/html,application/xhtml+xml,application/xml;q=0.9,"
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,*/*;q=0.8"
        ),
        "Accept-Language": "en-US,en;q=0.9",
    })
    for k, v in DISCLAIMER_COOKIES.items():
        s.cookies.set(k, v, domain="vaneck.com")
    return s


def _parse_money(text) -> Optional[float]:
    """Parse '$7,282,477.80' or '$.16' or '$-42,524.48' -> float USD."""
    if text is None:
        return None
    s = str(text).strip().replace(",", "").replace("$", "")
    if not s or s == "--":
        return None
    # The XLSX has values like "$.16" (no leading 0) — float() handles that fine.
    try:
        return float(s)
    except ValueError:
        return None


def _parse_int(text) -> Optional[int]:
    if text is None:
        return None
    s = str(text).strip().replace(",", "")
    if not s or s == "--":
        return None
    try:
        # Some rows have e.g. "1,991" stored as a string; openpyxl may also yield ints directly.
        return int(float(s))
    except ValueError:
        return None


def _parse_pct(text) -> Optional[float]:
    """Parse '9.45%' -> 9.45 (NOT 0.0945)."""
    if text is None:
        return None
    s = str(text).strip().replace("%", "")
    if not s or s == "--":
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _classify_row(ticker: str, asset_class: str) -> tuple[bool, bool]:
    """Return (is_cash, is_vaneck_fund) classification for a row."""
    t = (ticker or "").strip().upper()
    ac = (asset_class or "").strip().lower()
    # Cash positions: rows whose ticker is like "-USD CASH-" or asset_class contains "cash"
    is_cash = (
        "cash" in ac
        or t.startswith("-") and t.endswith("-") and "CASH" in t
        or t in {"--", ""}
    )
    is_vaneck_fund = t in VANECK_FUND_TICKERS
    return is_cash, is_vaneck_fund


def _parse_xlsx(xlsx_bytes: bytes) -> tuple[str, str, list[dict]]:
    """Parse the holdings XLSX. Returns (as_of_iso, sheet_name, holdings_rows)."""
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("XLSX is empty")

    # Row 0: title like "Daily Holdings (%)  05/12/2026"
    title = str(rows[0][0] or "")
    m = re.search(r"(\d{1,2})/(\d{1,2})/(\d{4})", title)
    if not m:
        raise ValueError(f"Could not find as-of date in title row: {title!r}")
    mo, day, yr = int(m.group(1)), int(m.group(2)), int(m.group(3))
    as_of = f"{yr:04d}-{mo:02d}-{day:02d}"

    # Find header row (contains 'Ticker')
    header_idx = None
    for i, row in enumerate(rows[:10]):
        if row and any(str(c or "").strip() == "Ticker" for c in row):
            header_idx = i
            break
    if header_idx is None:
        raise ValueError("Could not find header row in XLSX")

    header = [str(c or "").strip() for c in rows[header_idx]]
    col = {name: i for i, name in enumerate(header)}

    required = ["Ticker", "Holding Name", "Identifier (FIGI)", "Shares",
                "Asset Class", "Market Value (US$)", "% of Net Assets"]
    missing = [r for r in required if r not in col]
    if missing:
        raise ValueError(f"Missing expected columns: {missing}. Got header: {header}")

    holdings: list[dict] = []
    for row in rows[header_idx + 1:]:
        if not row or row[col["Ticker"]] in (None, ""):
            continue
        ticker_raw = str(row[col["Ticker"]] or "").strip()
        if not ticker_raw:
            continue
        # The footer disclaimer is a single string spanning column 0 — its other cells are None.
        # Skip rows where most cells are blank.
        if sum(1 for c in row if c not in (None, "")) <= 1:
            continue

        asset_class = str(row[col["Asset Class"]] or "").strip()
        is_cash, is_vaneck_fund = _classify_row(ticker_raw, asset_class)

        holding = {
            "ticker": ticker_raw,
            "name": str(row[col["Holding Name"]] or "").strip() or None,
            "figi": str(row[col["Identifier (FIGI)"]] or "").strip() or None,
            "shares": _parse_int(row[col["Shares"]]),
            "market_value_usd": _parse_money(row[col["Market Value (US$)"]]),
            "weight_pct": _parse_pct(row[col["% of Net Assets"]]),
            "asset_class": asset_class or None,
            "is_cash": is_cash,
            "is_vaneck_fund": is_vaneck_fund,
        }
        holdings.append(holding)

    return as_of, sheet_name, holdings


def _fetch_fund_metadata(session: requests.Session) -> dict:
    """Pull NAV / AUM / YTD return from the JSON dataset endpoint.

    Best-effort: if this endpoint fails we still have the XLSX, and the caller
    will fall back to deriving total_net_assets from sum(market_value).
    """
    try:
        r = session.get(FUND_DATASET_URL, timeout=15)
        r.raise_for_status()
        payload = r.json()
    except Exception:
        return {}

    def _money_to_usd(s):
        """e.g. '$75.72M' -> 75_720_000.0"""
        if not s:
            return None
        s = str(s).replace("$", "").replace(",", "").strip()
        m = re.match(r"([0-9.]+)\s*([KMBT])?$", s, re.IGNORECASE)
        if not m:
            return None
        val = float(m.group(1))
        mult = {"K": 1e3, "M": 1e6, "B": 1e9, "T": 1e12}.get((m.group(2) or "").upper(), 1)
        return val * mult

    def _pct_to_num(s):
        if not s:
            return None
        try:
            return float(str(s).replace("%", "").strip())
        except ValueError:
            return None

    def _money_simple(s):
        if not s:
            return None
        try:
            return float(str(s).replace("$", "").replace(",", "").strip())
        except ValueError:
            return None

    # Pull the most recent inception/update timestamp from holdings if present
    inception = payload.get("Inception Date")
    if inception:
        # Normalize "05/13/2025" -> "2025-05-13"
        m = re.match(r"(\d{1,2})/(\d{1,2})/(\d{4})", str(inception))
        if m:
            mo, day, yr = int(m.group(1)), int(m.group(2)), int(m.group(3))
            inception = f"{yr:04d}-{mo:02d}-{day:02d}"

    return {
        "nav": _money_simple(payload.get("NAV")),
        "ytd_return_pct": _pct_to_num(payload.get("YTD RETURNS")),
        "total_net_assets_usd": _money_to_usd(payload.get("Total Net Assets")),
        "gross_expense_ratio_pct": _pct_to_num(payload.get("Gross Expense Ratio")),
        "net_expense_ratio_pct": _pct_to_num(payload.get("Net Expense Ratio")),
        "inception_date": inception,
    }


def fetch() -> FetchResult:
    """Fetch and parse today's NODE holdings.

    Raises an exception on any failure — the caller decides whether to retry,
    log to stderr, or write a partial snapshot.
    """
    s = _session()

    # Visit the holdings page first to seed any session cookies the CDN sets.
    s.get(HOLDINGS_PAGE_URL, timeout=15)

    # Download the XLSX.
    r = s.get(
        HOLDINGS_DOWNLOAD_URL,
        timeout=20,
        headers={"Referer": HOLDINGS_PAGE_URL},
    )
    r.raise_for_status()
    if "spreadsheetml" not in r.headers.get("Content-Type", ""):
        raise RuntimeError(
            f"Expected an XLSX response, got Content-Type={r.headers.get('Content-Type')!r}"
        )

    # Extract filename from Content-Disposition (e.g. NODE_asof_20260512.xlsx).
    cd = r.headers.get("Content-Disposition", "")
    fn_match = re.search(r'filename="?(?P<name>[^";]+)"?', cd)
    source_filename = fn_match.group("name") if fn_match else "NODE_holdings.xlsx"

    # Parse the workbook.
    as_of, _sheet, holdings = _parse_xlsx(r.content)

    # Pull fund-level metadata (NAV, AUM, etc.) — non-fatal if missing.
    fund = _fetch_fund_metadata(s)

    # If we couldn't get AUM from the JSON endpoint, fall back to sum(market_value).
    if not fund.get("total_net_assets_usd"):
        total_mv = sum(h["market_value_usd"] or 0 for h in holdings)
        fund["total_net_assets_usd"] = round(total_mv, 2)

    fund["num_holdings"] = sum(
        1 for h in holdings if not h["is_cash"]
    )
    fund["num_cash_positions"] = sum(1 for h in holdings if h["is_cash"])

    return FetchResult(
        as_of=as_of,
        source_filename=source_filename,
        fetched_at=datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        fund=fund,
        holdings=holdings,
        raw_xlsx_bytes=r.content,
    )


if __name__ == "__main__":
    # Smoke test: print as-of + a few rows.
    result = fetch()
    print(f"As of: {result.as_of}")
    print(f"Source: {result.source_filename}")
    print(f"Fund: NAV=${result.fund.get('nav')}, "
          f"AUM=${result.fund.get('total_net_assets_usd'):,.0f}, "
          f"YTD={result.fund.get('ytd_return_pct')}%")
    print(f"Holdings: {len(result.holdings)} rows "
          f"({result.fund['num_holdings']} positions + "
          f"{result.fund['num_cash_positions']} cash)")
    print()
    for h in result.holdings[:5]:
        print(f"  {h['ticker']:<10} {h['name']:<35} "
              f"shares={h['shares']:>10,} "
              f"weight={h['weight_pct']:>5.2f}%")
